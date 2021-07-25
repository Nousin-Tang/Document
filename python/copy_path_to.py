# coding = utf-8
import os
import shutil
from openpyxl import load_workbook
from time import strftime, localtime

# 当前文件夹
cur_dir = os.getcwd()
# 获取项目根目录
project_root = os.path.dirname(os.path.realpath(__file__))

virtual_drive = 'z:'
# 是否是本地资源
local_source_flag = True
# 源文件夹
source_path = 'F:\\py\\product'
# 目标文件夹
target_path = cur_dir + os.sep + 'temp'
# 删除上一次执行结果
delete_pre_result = False
# -------------------------------------
# |  货号  |  品名 | 小类名称 | 大类名称 |
# -------------------------------------
# | xxxxxx | xxxx |  xxxxx  | xxxxxxxx |
# -------------------------------------
excel_path = 'excel.xlsx'

# 获取原文件夹内容
file_path_list_for_search = []

cmd_ls_dir_file_path = cur_dir + os.sep + 'folder.txt'


# 初始化 执行的方法
def init():
    print('初始化...')
    # 创建文输出文件夹
    if not os.path.exists(target_path):
        os.makedirs(target_path)
    elif delete_pre_result:
        # 删除上一次执行结果
        del_list = os.listdir(target_path)
        for f_ in del_list:
            file_path = os.path.join(target_path, f_)
            if os.path.isfile(file_path):
                os.remove(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
    if local_source_flag:
        # 本地文件
        os.system('dir ' + source_path + ' /AD /s /O /B > ' + cmd_ls_dir_file_path)
        print('%s 成功列出源文件夹下的文件路径到指定文件' % strftime("%Y-%m-%d %H:%M:%S", localtime()))
    else:
        # 获取共享资源指定目录: net use k: \\xxx.xxx.xxx.xxx\D\360 将 xxx.xxx.xxx.xxx 的共享资源下的 D 盘符下 360 目录影射到本地
        if not os.path.exists(virtual_drive):
            print('创建虚拟路径：%s' % virtual_drive)
            os.system("net use " + virtual_drive + " \\" + source_path)

        # 列出源文件夹下的文件路径到指定文件(/a:-d 只列出文件)
        print('%s 列出源文件夹下的文件路径到指定文件：%s' % (strftime("%Y-%m-%d %H:%M:%S", localtime()), cmd_ls_dir_file_path))
        # 列出所有文件 dir  /s /b /a:-d > 1.txt
        # os.system('dir ' + virtual_drive + ' /s /b /a:-d > ' + cmd_ls_dir_file_path)
        # 列出所有文件夹 dir /AD /s /O /B > 目录名.txt
        os.system('dir ' + virtual_drive + ' /AD /s /O /B > ' + cmd_ls_dir_file_path)
        print('%s 成功列出源文件夹下的文件路径到指定文件' % strftime("%Y-%m-%d %H:%M:%S", localtime()))


# Excel 中文件内容
class ExcelContent:
    def __init__(self, product_no, small_category, large_category):
        # 货号
        self.product_no = product_no
        # 小类名称
        self.small_category = small_category
        # 大类名称
        self.large_category = large_category


# -------------------------------------
# |  货号  |  品名 | 小类名称 | 大类名称 |
# -------------------------------------
# | xxxxxx | xxxx |  xxxxx  | xxxxxxxx |
# -------------------------------------
# 读取 Excel 中内容
def read_excel():
    # 文件路径
    file_path = os.path.join(project_root, excel_path)
    # 打开一个workbook
    wb = load_workbook(filename=file_path, data_only=True)
    sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
    # values 用来存放数据
    values = []
    for rowNum in range(3, sheet.max_row + 1):
        product_no = str(sheet.cell(row=rowNum, column=1).value)
        small_category = str(sheet.cell(row=rowNum, column=3).value)
        large_category = str(sheet.cell(row=rowNum, column=4).value)
        values.append(ExcelContent(product_no, small_category, large_category))
    # 将table中第一行的数据读取并添加到data_list中
    return values


# Excel 中文件内容
class FileContent:
    def __init__(self, path, sub_path):
        # 目标文件夹
        self.path = path
        # 包含路径
        self.sub_path = sub_path


# 匹配目标文件夹
def match_path(files):
    result = []
    # 读取文件并缓存
    if len(file_path_list_for_search) == 0:
        print('... 读取已缓存的文件路径')
        with open(cmd_ls_dir_file_path) as f_:
            for line in f_.readlines():
                line_str = str(line).strip()
                if line_str == 'None' or line_str == '':
                    continue
                file_path_list_for_search.append(line_str)
        print('... 扫描到 %d 个文件夹' % len(file_path_list_for_search))

    for f in file_path_list_for_search:
        for s in files:
            if f.endswith(os.sep + s.product_no):
                result.append(FileContent(f, s.large_category + os.sep + s.small_category + os.sep + s.product_no))
                break
    return result


# 移动文件夹
def transfer_file(source_file_contents):
    for file_content in source_file_contents:
        src = os.path.join(file_content.path)
        dst = os.path.join(target_path, file_content.sub_path)
        # if not os.path.exists(dst):
        #     os.makedirs(dst)
        # print(os.listdir(src))
        # 拷贝文件夹
        shutil.copytree(src, dst, dirs_exist_ok=True)
        # 移动文件
        # shutil.copy(src, dst)
        # 移动文件（夹）
        # shutil.move(src, dst)
        # 删除文件夹及内容
        # shutil.rmtree(src)


# 拷贝文件夹
def copy_dirs(src, dst):
    if not os.path.exists(dst):  # 如不存在目标目录则创建
        os.makedirs(dst)
    files = os.listdir(src)  # 获取文件夹中文件和目录列表
    for f in files:
        if os.path.isdir(src + os.sep + f):  # 判断是否是文件夹
            copy_dirs(src + os.sep + f, dst + os.sep + f)  # 递归调用本函数
        else:
            shutil.copyfile(src + os.sep + f, dst + os.sep + f)  # 拷贝文件


if __name__ == '__main__':
    try:
        init()
        file_path_list = read_excel()
        product_nos = len(file_path_list)
        print('读取到 %d 个货号' % product_nos)
        if product_nos > 0:
            print('查找匹配结果...')
            source_result = match_path(file_path_list)
            print('... 读取到 %d 个匹配记录' % len(source_result))
            transfer_file(source_result)
            print('文件夹传输完成...')
    finally:
        if os.path.exists(virtual_drive):
            # 删除共享资源链接: net use \\xxx.xxx.xxx.xxx /delete
            os.system("net use " + virtual_drive + " /delete")
            print('删除临时驱动盘 %s' % virtual_drive)
