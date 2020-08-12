# coding=utf-8
# xlsx 格式的 Excel 读写
import os

# xls 格式的 Excel 写入
import xlwt
# excel文件的读、写和修改，只能处理xlsx文件
from openpyxl import load_workbook, Workbook
from openpyxl.comments import Comment
from ctypes import *

# 导入：python -m pip install pypiwin32
import win32com.client as win32

# 工程目录(E:\project\NewRetailing\01 Development\03 Code\python)
PROJECT_ROOT = os.path.dirname(os.path.realpath(__file__))

# E:\project\NewRetailing\01 Development
root_path = os.path.dirname(os.path.dirname(os.path.dirname(os.path.realpath(__file__))))
# E:\project\NewRetailing\01 Development\01 Requirements\国际化翻译
source_path = root_path + os.sep + "01 Requirements" + os.sep + "国际化翻译"
# E:\project\NewRetailing\01 Development\01 Requirements\国际化翻译\excel_template
target_path = source_path + os.sep + "excel_template"

# 模板文件名称多语言文件
excel_name_file_path = source_path + os.sep + '9-Dict字典表-翻译.xlsx'
# 模板内容多语言文件
excel_content_file_path = source_path + os.sep + '8-页面导入导出Excel模板-翻译.xlsx'
# 生成 Excel 模板存放位置
excel_template_save_path = target_path

# 中文
zh_key = 0
# 英语
en_key = 1
# 阿语
ar_key = 2
# 法语
fr_key = 3
# 俄语
ru_key = 4
# 西语
es_key = 5
# 斯里兰卡
lk_key = 6
# 越语
vi_key = 7

# 语言简称
zh_ = 'zh'
en_ = 'en'
ar_ = 'ar'
fr_ = 'fr'
ru_ = 'ru'
es_ = 'es'
lk_ = 'lk'
vi_ = 'vi'

# 语言简称集合(生成的模板 按照给定的语言生成)
# lang_names = [zh_, en_, ar_, fr_, ru_, es_]
lang_names = [zh_]

# 是否删除旧文件
delete_old_file = False

# 系统名称
sys_marketing = 'marketing'
sys_store = 'store'
sys_ams = 'ams'
sys_base = 'base'
sys_wbs = 'wbs'

# 系统名称集合
systems = (sys_marketing, sys_store, sys_ams, sys_base, sys_wbs)

# 生成的文件路径
generate_file_path = []


# Excel 名称，类型，系统
# （多语言数据要与 lang_names 数据一致）
class ExcelName:

    def __init__(self, system, exl_type, zh, en, es, ar, fr, ru):
        # 系统（marketing, store, ams, base, wbs）
        self.system = system
        # 模板类型：1-导入模板，2-导出模板
        self.exl_type = exl_type
        # 多语言名称
        self.zh = zh
        self.en = en
        self.es = es
        self.ar = ar
        self.fr = fr
        self.ru = ru


# Excel 内容
# （多语言数据要与 lang_names 数据一致）
class ExcelContent:

    def __init__(self, excel_name, comment, export_content, zh, en, es, ar, fr, ru):
        # 模板名称
        self.excel_name = excel_name
        # 备注
        self.comment = comment
        # 导出模板 填充内容
        self.export_content = export_content
        # 多语言 title
        self.zh = zh
        self.en = en
        self.es = es
        self.ar = ar
        self.fr = fr
        self.ru = ru


# Excel 列
class ExcelColumn:

    def __init__(self, content, comment, export_temp_content):
        # 内容
        self.content = content
        # 备注
        self.comment = comment
        # 导出模板内筒
        self.export_temp_content = export_temp_content


def resolve_excel_name_and_extension(excel_name):
    excel_name_ = excel_name.replace('\\', '')
    return excel_name_[0:excel_name_.find('.')] + '.xlsx'


# 读取 Excel 多语言模板名称
def read_excel_name():
    wb = load_workbook(filename=excel_name_file_path, data_only=True)
    # 获取 Sheet
    sheet = wb['翻译']
    result = []
    for rowNum in range(3, sheet.max_row + 1):
        excel_type = str(sheet.cell(row=rowNum, column=2).value)
        if excel_type != 'excel_template':
            continue

        description = sheet.cell(row=rowNum, column=4).value
        if description is None:
            continue

        zh_col = sheet.cell(row=rowNum, column=5).value
        en_col = sheet.cell(row=rowNum, column=7).value
        es_col = sheet.cell(row=rowNum, column=8).value
        ar_col = sheet.cell(row=rowNum, column=9).value
        fr_col = sheet.cell(row=rowNum, column=10).value
        ru_col = sheet.cell(row=rowNum, column=11).value

        description = str(description)
        zh = '' if zh_col is None else resolve_excel_name_and_extension(str(zh_col))
        en = '' if en_col is None else resolve_excel_name_and_extension(str(en_col))
        es = '' if es_col is None else resolve_excel_name_and_extension(str(es_col))
        ar = '' if ar_col is None else resolve_excel_name_and_extension(str(ar_col))
        fr = '' if fr_col is None else resolve_excel_name_and_extension(str(fr_col))
        ru = '' if ru_col is None else resolve_excel_name_and_extension(str(ru_col))

        for sys_name in systems:
            if description.find(sys_name) >= 0:
                result.append(ExcelName(sys_name, 1 if description.find('导入') >= 0 else 2, zh, en, es, ar, fr, ru))
    return result


# 读取 Excel 中的内容
def read_excel_content():
    wb = load_workbook(filename=excel_content_file_path, data_only=True)
    # 获取表索引
    sheet = wb['翻译']
    result = []
    for rowNum in range(3, sheet.max_row + 1):
        excel_name = sheet.cell(row=rowNum, column=2).value
        comment = sheet.cell(row=rowNum, column=3).comment
        zh_col = sheet.cell(row=rowNum, column=3).value
        export_content = sheet.cell(row=rowNum, column=4).value
        en_col = sheet.cell(row=rowNum, column=5).value
        es_col = sheet.cell(row=rowNum, column=6).value
        ar_col = sheet.cell(row=rowNum, column=7).value
        fr_col = sheet.cell(row=rowNum, column=8).value
        ru_col = sheet.cell(row=rowNum, column=9).value

        name = '' if excel_name is None else str(excel_name)
        col_comment = '' if comment is None else str(comment.content)
        exp_cont = '' if export_content is None else str(export_content)
        zh = '' if zh_col is None else str(zh_col)
        en = '' if en_col is None else str(en_col)
        es = '' if es_col is None else str(es_col)
        ar = '' if ar_col is None else str(ar_col)
        fr = '' if fr_col is None else str(fr_col)
        ru = '' if ru_col is None else str(ru_col)

        result.append(ExcelContent(name, col_comment, exp_cont, zh, en, es, ar, fr, ru))
    return result


# 获取 Excel 多语言内容（title, comment）等信息
def get_excel_content(contents, excel_name):
    result = []
    for con in contents:
        name = excel_name.zh[0:excel_name.zh.find(".")]
        name1 = con.excel_name[0:con.excel_name.find(".")]
        if name == name1:
            result.append(con)
    return result


# 创建 xlsx 格式的 Excel
def create_xlsx(system, exl_type, exl_name, cols, lang):
    if exl_name is None or exl_name == '':
        return
    wb = Workbook()
    sheet = wb.create_sheet('data', 0)
    for i in range(0, len(cols)):
        col = cols[i]

        sheet.cell(row=1, column=i + 1).value = col.content
        if col.comment != '':
            sheet.cell(row=1, column=i + 1).comment = Comment(col.comment, '')
        if exl_type == 2:
            sheet.cell(row=2, column=i + 1).value = col.export_temp_content

    file_path = excel_template_save_path + os.sep + system + os.sep + lang
    # 创建文件夹
    if not os.path.exists(file_path):
        os.makedirs(file_path)

    # 保存 Excel 模板
    _file_path_ = file_path + os.sep + exl_name
    # 文件是否存在
    file_exists = os.path.exists(_file_path_)
    if file_exists and delete_old_file:
        os.remove(_file_path_)

    # 如果文件存在且不删除 则返回
    if file_exists and not delete_old_file:
        return
    # 创建
    wb.save(_file_path_)
    generate_file_path.append(_file_path_)


# 创建 xls 格式的 Excel
def create_xls(system, exl_type, exl_name, cols, lang):
    if exl_name is None or exl_name == '':
        return
    wb = xlwt.Workbook(encoding='utf-8')
    sheet = wb.add_sheet('data', cell_overwrite_ok=True)
    for i in range(0, len(cols)):
        col = cols[i]
        if col.comment != '':
            sheet.write(0, i, col.content)
        else:
            sheet.write(0, i, col.content)
        if exl_type == 2:
            sheet.write(1, i, col.export_temp_content)

    file_path = excel_template_save_path + os.sep + system + os.sep + lang
    # 创建文件夹
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    # 保存 Excel 模板
    _file_path_ = file_path + os.sep + exl_name
    # 文件是否存在
    file_exists = os.path.exists(_file_path_)
    if file_exists and delete_old_file:
        os.remove(_file_path_)

    # 如果文件存在且不删除 则返回
    if file_exists and not delete_old_file:
        return
    # 创建
    wb.save(_file_path_)
    generate_file_path.append(_file_path_)


# 生成 Excel 模板
def generate_excel(excel_name, excel_content):
    # 处理中文模板
    exl_type = excel_name.exl_type
    system = excel_name.system

    for lang in lang_names:
        exl_name = getattr(excel_name, lang)
        exl_cols = []
        for content in excel_content:
            getattr(content, lang)
            exl_cols.append(ExcelColumn(getattr(content, lang), content.comment, content.export_content))
        if len(exl_cols) > 0:
            # create_xls(system, exl_type, exl_name, exl_cols, lang)
            create_xlsx(system, exl_type, exl_name, exl_cols, lang)


def change_xlsx_to_xls():
    excel = win32.gencache.EnsureDispatch('excel.application')
    excel.DisplayAlerts = 0

    for file in generate_file_path:
        if os.path.isdir(file):
            return
        file_name = os.path.splitext(file)  # 文件和格式分开
        if file_name[1] == '.xlsx':
            transfer_file1 = file  # 要转换的excel
            transfer_file2 = file_name[0]  # 转换出来excel

            pro = excel.Workbooks.Open(transfer_file1)  # 打开要转换的excel
            pro.SaveAs(transfer_file2 + ".xls", FileFormat=56)  # 另存为xls格式
            pro.Close()
            if os.path.exists(transfer_file1):  # 如果文件存在
                # 删除文件，可使用以下两种方法。
                os.remove(transfer_file1)

    excel.Application.Quit()

# 读取数据 并生成模板
def read_data_and_generate_template():
    # 读取Excel模板名称
    names = read_excel_name()

    # 读取Excel模板内容信息
    contents = read_excel_content()

    for excel_name in names:
        excel_content = get_excel_content(contents, excel_name)
        if len(excel_content) > 0:
            generate_excel(excel_name, excel_content)
    # 转换 Excel
    change_xlsx_to_xls()


# 转换 xlsx 格式的模板为 xls
def transfer_xlsx_to_xls(file):
    if os.path.isdir(file):
        return
    file_name = os.path.splitext(file)  # 文件和格式分开
    if file_name[1] == '.xlsx':
        transfer_file1 = file  # 要转换的excel
        transfer_file2 = file_name[0]  # 转换出来excel
        excel = win32.gencache.EnsureDispatch('excel.application')
        excel.DisplayAlerts = 0
        pro = excel.Workbooks.Open(transfer_file1)  # 打开要转换的excel
        pro.SaveAs(transfer_file2 + ".xls", FileFormat=56)  # 另存为xls格式
        pro.Close()
        excel.Application.Quit()


# 遍历文件夹
def walk_file(file):
    # root 表示当前正在访问的文件夹路径
    # dirs 表示该文件夹下的子目录名list
    # files 表示该文件夹下的文件list
    for root, dirs, files in os.walk(file):
        # 遍历文件
        for f in files:
            transfer_xlsx_to_xls(os.path.join(root, f))

        # 遍历所有的文件夹
        for d in dirs:
            walk_file(os.path.join(root, d))


if __name__ == '__main__':
    # 处理生成 xlsx 格式模板
    read_data_and_generate_template()
    # 转换 xlsx 格式的模板为 xls
    # walk_file(target_path)
