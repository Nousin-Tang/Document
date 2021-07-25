# coding=utf-8
from openpyxl import load_workbook
import time
import os

FROM_EXCEL_NAME = 'excel.xlsx'
# 获取项目根目录
PROJECT_ROOT = os.path.dirname(os.path.realpath(__file__))
# 文件路径
filePath = os.path.join(PROJECT_ROOT, FROM_EXCEL_NAME)


# 读取excel的数据
def read_excel(file_path):
    # 打开一个workbook
    wb = load_workbook(filename=file_path, data_only=True)
    sheet_names = wb.sheetnames
    # 需要生成创建语句的表
    _list = list()

    for i in range(0, len(sheet_names)):
        # 获取表索引
        index_sheet = wb[sheet_names[i]]

        for rowNum in range(2, index_sheet.max_row + 1):
            value_6 = index_sheet.cell(row=rowNum, column=6).value
            if value_6 is not None:
                _list.append(value_6)

    def wordcount(_list):
        dict1 = {}
        for every_world in _list:
            if every_world in dict1:
                dict1[every_world] += 1
            else:
                dict1[every_world] = 1
        return dict1

    dict_count = wordcount(_list)

    print(dict_count)

    # # 当前时间
    # datetime = time.strftime('%Y-%m-%d %H%M%S', time.localtime(time.time()))
    # sql_file_name = filename[0:find_last(filename, '.')] + datetime + '.xls'
    # # 打开文档，没有则创建
    # sql_file = open(sql_file_name, 'w')
    # sql_file.write("\n\n".join(table_sql_list))
    # sql_file.close()

    def save_to_excel(data, sheet_name, wb_name):
        print("写入excel：")
        wb_ = load_workbook(filename=wb_name, read_only=False)
        ws = wb_.create_sheet(sheet_name)
        for key in data.keys():
            ws.append([key, data[key]])

        wb_.save(filename=wb_name)  # 保存数据
        print("保存成功")

    save_to_excel(dict_count, "result", file_path)


if __name__ == '__main__':
    read_excel(filePath)
