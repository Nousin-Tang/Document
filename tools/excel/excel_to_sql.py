# coding=utf-8
from openpyxl import load_workbook
import time


# 查找字符字符串最后一次出现的位置
def find_last(_string, _str):
    last_position = -1
    while True:
        position = _string.find(_str, last_position + 1)
        if position == -1:
            return last_position
        last_position = position


# 读取excel的数据
def read_excel():
    filename = '/Users/unnous/My/DB-Design.xlsx'
    # 打开一个workbook
    wb = load_workbook(filename=filename, data_only=True)
    # 获取表索引
    index_sheet = wb.get_sheet_by_name('index')
    # 需要生成创建语句的表
    table_list = list()
    for rowNum in range(2, index_sheet.max_row + 1):
        value_6 = index_sheet.cell(row=rowNum, column=6).value
        if value_6 is not None and value_6 == 'Y':
            table_index = index_sheet.cell(row=rowNum, column=2).value
            if table_index is not None:
                table_list.append(table_index)

    if len(table_list) < 1:
        return

    for table_index_name in table_list:
        sheet = wb.get_sheet_by_name(table_index_name)
        table_name = str(sheet['b2'].value)
        table_comment = str(sheet['e2'].value)
        table_sql = 'CREATE TABLE ' + table_name + ' ('
        primary_key_set = list()
        index_sql = ''
        unique_key_set = list()
        for rowNum in range(5, sheet.max_row + 1):
            column_name = str(sheet.cell(row=rowNum, column=1).value)
            # TINYINT,INT,BIGINT,DECIMAL,CHAR,VARCHAR,TEXT,LONGTEXT,BLOB,DATE,DATETIME,TIMESTAMP
            column_type = str(sheet.cell(row=rowNum, column=2).value)
            column_length = str(sheet.cell(row=rowNum, column=3).value)
            column_can_null = str(sheet.cell(row=rowNum, column=4).value)
            column_default = str(sheet.cell(row=rowNum, column=5).value)
            column_index = str(sheet.cell(row=rowNum, column=6).value)
            column_desc = str(sheet.cell(row=rowNum, column=7).value)

            table_sql += '\n\t' + column_name + ' ' + column_type

            # 长度
            column_type_upper = column_type.upper()
            if column_type_upper == 'TINYINT':
                if column_length != 'None':
                    table_sql += '(' + column_length + ')'
                else:
                    table_sql += '(2)'

            elif column_type_upper == 'INT':
                if column_length != 'None':
                    table_sql += '(' + column_length + ')'

            elif column_type_upper == 'BIGINT':
                if column_length != 'None':
                    table_sql += '(' + column_length + ')'

            elif column_type_upper == 'DECIMAL':
                if column_length != 'None':
                    table_sql += '(' + column_length + ')'
                else:
                    table_sql += '(18，5)'

            elif column_type_upper == 'CHAR':
                if column_length != 'None':
                    table_sql += '(' + column_length + ')'
                else:
                    table_sql += '(10)'

            elif column_type_upper == 'VARCHAR':
                if column_length != 'None':
                    table_sql += '(' + column_length + ')'
                else:
                    table_sql += '(255)'

            # 是否可空
            if column_can_null != 'None' and column_can_null.upper() == 'N':
                table_sql += ' NOT NULL'

            # 默认值
            if column_default != 'None':
                default = column_default
                if 'AUTO' == default.upper():
                    table_sql += ' AUTO_INCREMENT'
                elif 'CURRENT' == default.upper():
                    table_sql += ' DEFAULT CURRENT_TIMESTAMP'
                elif column_type_upper == 'TINYINT' \
                        or column_type_upper == 'INT' \
                        or column_type_upper == 'BIGINT' \
                        or column_type_upper == 'DECIMAL' \
                        or column_type_upper == 'DATETIME' \
                        or column_type_upper == 'TIMESTAMP':
                    table_sql += ' DEFAULT ' + default
                elif column_type_upper == 'CHAR' \
                        or column_type_upper == 'VARCHAR' \
                        or column_type_upper == 'TEXT' \
                        or column_type_upper == 'LONGTEXT':
                    table_sql += ' DEFAULT \'' + default + '\''
                else:
                    table_sql += ' DEFAULT ' + default
            # 备注
            if column_desc != 'None':
                table_sql += " COMMENT '" + column_desc + "'"

            # 主键与索引判断
            if column_index != 'None':
                index_upper = column_index.upper()
                if 'PRIMARY' == index_upper:
                    primary_key_set.append(column_name)
                elif 'INDEX' == index_upper:
                    index_sql += '\nALTER TABLE ' + table_name + \
                                 ' ADD INDEX index_' + column_name + ' (' + column_name + ');'
                elif 'UNIQUE' == index_upper:
                    unique_key_set.append(column_name)

            if rowNum != sheet.max_row:
                table_sql += ","
            else:
                if primary_key_set:
                    table_sql += ',\n\tPRIMARY KEY (' + ','.join(primary_key_set) + ') USING BTREE'

        table_sql += '\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 ROW_FORMAT=DYNAMIC COMMENT=\'' + table_comment + '\';'
        table_sql += index_sql
        table_sql += '\nALTER TABLE ' + table_name + ' ADD UNIQUE (' + ','.join(unique_key_set) + ');'

        # 当前时间
        datetime = time.strftime('%Y-%m-%d %H%M%S', time.localtime(time.time()))
        sql_file_name = filename[0:find_last(filename, '.')] + datetime + '.sql'
        # 打开文档，没有则创建
        sql_file = open(sql_file_name, 'w')
        sql_file.write(table_sql)
        sql_file.close()


if __name__ == '__main__':
    read_excel()
